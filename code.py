import pandas as pd
import requests
from datetime import datetime, timedelta, timezone
import openpyxl
from openpyxl import load_workbook
import os

excel_path = r"C:\\Users\selim.olmus\\OneDrive - Horoz Lojistik\\_slms_\\Opet_Yakıt\\Data2.xlsx"

# ---- BÖLÜM 1: VERİ ÇEKME VE GÜNCELLEME ----

# 1. Durum sekmesini oku (DistrictCode'u string olarak al!)
df_durum = pd.read_excel(excel_path, sheet_name="durum", dtype={"DistrictCode": str})

# 2. EndDate = yarın
tomorrow = (datetime.now(timezone.utc) + timedelta(days=1)).strftime('%Y-%m-%dT00:00:00.0Z')

# 3. Excel dosyasını yükle
wb = openpyxl.load_workbook(excel_path)

for index, row in df_durum.iterrows():
    district_code = row['DistrictCode']
    start_date = pd.to_datetime(row['StartDate']).strftime('%Y-%m-%dT00:00:00.0Z')

    # API URL'si
    url = (
        f"https://api.opet.com.tr/api/fuelprices/prices/archive"
        f"?DistrictCode={district_code}&StartDate={start_date}"
        f"&EndDate={tomorrow}&IncludeAllProducts=true"
    )

    # 4. API çağrısı
    response = requests.get(url)
    if response.status_code != 200:
        continue

    # 5. Gelen veriyi işle
    try:
        data = response.json()
    except ValueError:
        continue

    records = []

    for day_item in data:
        for price in day_item.get("prices", []):
            if price.get("productName") == "Motorin UltraForce":
                records.append({
                    "priceDate": price.get("priceDate"),
                    "productName": price.get("productName"),
                    "amount": price.get("amount")
                })

    if not records:
        continue

    df_new = pd.DataFrame(records)

    # 6. Tarih dönüştürme ve zaman dilimini silme
    df_new["priceDate"] = pd.to_datetime(df_new["priceDate"], errors="coerce").dt.tz_localize(None)

    sheet_name = str(district_code)

    try:
        df_old = pd.read_excel(excel_path, sheet_name=sheet_name)
        if df_old.empty:
            df_combined = df_new
        else:
            df_combined = pd.concat([df_old, df_new], ignore_index=True).drop_duplicates()
    except:
        df_combined = df_new

    # 7. Sayfaya yaz
    with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        df_combined.to_excel(writer, sheet_name=sheet_name, index=False)

    # 8. StartDate hücresini güncelle
    valid_dates = df_combined["priceDate"].dropna()
    if not valid_dates.empty:
        latest_date = valid_dates.max().strftime('%Y-%m-%d')
        df_durum.at[index, "StartDate"] = latest_date

# 9. Durum sekmesini güncelle
with pd.ExcelWriter(excel_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
    df_durum.to_excel(writer, sheet_name="durum", index=False)

# ---- BÖLÜM 2: ESKALASYON HESAPLAMA ----

# Tüm sonuçları saklamak için boş bir DataFrame oluştur
tum_sonuclar = pd.DataFrame()

# Hesapla ve ekle fonksiyonu
def hesapla_ve_ekle(data, rate):
    if data.empty:
        return data
    
    # Önce NaN değerlerini temizle
    data = data.dropna(subset=['priceDate', 'amount'])
    
    # Tarihe göre sırala
    data = data.sort_values(by='priceDate')
    
    # Değişim ve eskalasyon sütunlarını oluştur
    data["degisim"] = 0.0
    data["eskalasyon"] = 0.0
    
    # İlk satır için referans değeri al
    onceki_tutar = data.iloc[0]["amount"]
    
    # İlk satır için değişim 0 olarak ayarla
    for i in range(1, len(data)):  # 1'den başla, ilk satır için değişim hesaplanmaz
        mevcut_tutar = data.iloc[i]["amount"]
        if pd.notna(mevcut_tutar) and pd.notna(onceki_tutar) and onceki_tutar != 0:
            degisim = mevcut_tutar / onceki_tutar - 1
            data.iloc[i, data.columns.get_loc("degisim")] = degisim
            
            if abs(degisim) >= rate:
                onceki_tutar = mevcut_tutar
                data.iloc[i, data.columns.get_loc("eskalasyon")] = degisim
            else:
                data.iloc[i, data.columns.get_loc("eskalasyon")] = 0
    
    return data

# Hesapla sekmesi verisini al ve tarihi datetime'a çevir
try:
    hesapla_data = pd.read_excel(excel_path, sheet_name='hesapla')
    hesapla_data['priceDate'] = pd.to_datetime(hesapla_data['priceDate'])

    # Her satır için işlemi tekrarlayalım
    for index, row in hesapla_data.iterrows():
        district_code = str(row["DistrictCode"]).zfill(6)
        name = row["Name"]
        
        # Rate değerini al (yoksa varsayılan olarak 0.05 kullan)
        rate = row.get("rate", 0.05)
        if pd.isna(rate):
            rate = 0.05
        
        try:
            # Tarih dönüşümü
            current_price_date = pd.to_datetime(row["priceDate"])
            
            # Sayfa okuma
            district_data = pd.read_excel(excel_path, sheet_name=district_code)
            district_data['priceDate'] = pd.to_datetime(district_data['priceDate'], errors='coerce')
            
            # NaT değerlerini temizle
            district_data = district_data.dropna(subset=['priceDate'])
            
            # Filtreleme
            filtered_data = district_data[district_data['priceDate'] >= current_price_date].copy()
            
            if filtered_data.empty:
                continue
            
            # Değişimleri hesapla ve ekle (rate değerini de gönder)
            filtered_data = hesapla_ve_ekle(filtered_data, rate)
            
            # Name ve rate sütunlarını ekle
            filtered_data['Name'] = name
            filtered_data['rate'] = rate
            
            # Sonuçları tüm sonuçlar DataFrame'ine ekle
            tum_sonuclar = pd.concat([tum_sonuclar, filtered_data])
            
        except Exception:
            continue

    # İlk satırı ve eskalasyon değeri 0 olmayanları filtrele
    if not tum_sonuclar.empty:
        # Her bölge için ilk satırı al
        ilk_satirlar = tum_sonuclar.groupby('Name').first().reset_index()
        
        # Eskalasyon değeri 0 olmayan satırları al
        eskalasyon_satirlar = tum_sonuclar[tum_sonuclar['eskalasyon'] != 0]
        
        # İki DataFrame'i birleştir
        sonuc_df = pd.concat([ilk_satirlar, eskalasyon_satirlar]).drop_duplicates()
        
        # Tarihe göre sırala
        sonuc_df = sonuc_df.sort_values(by=['Name', 'priceDate'])
        
        # Excel'e yaz
        try:
            # Dosya varsa, mevcut dosyayı yükle
            if os.path.exists(excel_path):
                # Eğer 'eskalasyon' sayfası varsa, önce onu sil
                if 'eskalasyon' in pd.ExcelFile(excel_path).sheet_names:
                    # Mevcut dosyayı yükle
                    book = load_workbook(excel_path)
                    # Eskalasyon sayfasını sil
                    if 'eskalasyon' in book.sheetnames:
                        idx = book.sheetnames.index('eskalasyon')
                        book.remove(book.worksheets[idx])
                    # Değişiklikleri kaydet
                    book.save(excel_path)
                
                # Şimdi append modunda yeni sayfayı ekle
                with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
                    sonuc_df.to_excel(writer, sheet_name='eskalasyon', index=False)
            else:
                # Dosya yoksa yeni oluştur
                sonuc_df.to_excel(excel_path, sheet_name='eskalasyon', index=False)
        except Exception:
            pass
except Exception:
    pass
 # type: ignore
